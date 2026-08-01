"""Import vendor_master.xlsx / ownership_matrix.xlsx into SQLite (§10, §18).
Excel stays the human-edited source; DB is the runtime mirror. Re-runnable (upsert).
Usage: python scripts/import_registers.py [data_templates_dir]
"""
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from orchestrator.config import CFG
from orchestrator.db import get_db, init_db


def import_vendors(conn, xlsx: Path) -> int:
    ws = openpyxl.load_workbook(xlsx)["vendors"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    n = 0
    for r in rows:
        if not r or not r[0]:
            continue
        (name, tier, domains, family, contact, title, email, phone, country,
         auth, dealreg, role, owner, status, validated) = r[:15]
        email_domain = str(email).split("@")[-1].lower() if email else None
        conn.execute(
            """INSERT INTO vendors (vendor_name, tier, tech_domains, product_family,
               contact_name, job_title, email, email_domain, phone, country,
               vendor_authorised, deal_reg_capable, role, assigned_nl_owner,
               contact_status, last_validated)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(vendor_name, email) DO UPDATE SET
               tier=excluded.tier, tech_domains=excluded.tech_domains,
               contact_status=excluded.contact_status, last_validated=excluded.last_validated""",
            (name, tier, json.dumps([d.strip() for d in str(domains).split(";")]),
             family, contact, title, email, email_domain, phone, country or "UAE",
             int(auth or 0), int(dealreg or 0), role or "SALES", owner,
             status or "Unverified", str(validated) if validated else None),
        )
        n += 1
    return n


def import_ownership(conn, xlsx: Path) -> int:
    ws = openpyxl.load_workbook(xlsx)["ownership"]
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        domain, oem, family, primary, backup, commercial, reviewer, escalation = r[:8]
        conn.execute(
            """INSERT INTO ownership_matrix (tech_domain, oem, product_family,
               primary_owner, backup_owner, commercial_owner, technical_reviewer,
               escalation_manager) VALUES (?,?,?,?,?,?,?,?)
               ON CONFLICT(tech_domain, oem, product_family) DO UPDATE SET
               primary_owner=excluded.primary_owner, backup_owner=excluded.backup_owner,
               commercial_owner=excluded.commercial_owner,
               technical_reviewer=excluded.technical_reviewer,
               escalation_manager=excluded.escalation_manager""",
            (domain, oem, family, primary, backup, commercial, reviewer, escalation),
        )
        n += 1
    return n


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data_templates")
    init_db(CFG.db_path)
    conn = get_db(CFG.db_path)
    with conn:
        nv = import_vendors(conn, src / "vendor_master.xlsx")
        no = import_ownership(conn, src / "ownership_matrix.xlsx")
    print(f"imported {nv} vendors, {no} ownership rows")
